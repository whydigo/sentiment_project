import pandas as pd
import numpy as np
from tqdm import tqdm
import time
import os
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter as get_excel_letter
from openpyxl.utils import column_index_from_string

class ExcelProcessor:
    def __init__(self, sentiment_analyzer, topic_classifier):
        self.sentiment_analyzer = sentiment_analyzer
        self.topic_classifier = topic_classifier
        self.allowed_extensions = {'xlsx', 'xls', 'csv'}
        
        self.sentiment_colors = {
            'positive': '92D050',
            'negative': 'FF6B6B',
            'neutral': 'FFD966'
        }
    
    def get_column_letter(self, index):
        try:
            return get_excel_letter(index + 1)
        except:
            if index < 26:
                return chr(65 + index)
            else:
                first = chr(65 + (index // 26) - 1)
                second = chr(65 + (index % 26))
                return first + second
    
    def letter_to_index(self, letter):
        try:
            return column_index_from_string(letter) - 1
        except:
            letter = letter.upper()
            result = 0
            for i, char in enumerate(reversed(letter)):
                result += (ord(char) - 64) * (26 ** i)
            return result - 1
    
    def allowed_file(self, filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in self.allowed_extensions
    
    def read_excel(self, filepath, sheet_name=0, column_name=None, column_index=0):
        try:
            if filepath.endswith('.csv'):
                df = pd.read_csv(filepath, encoding='utf-8')
            else:
                df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')
            
            if column_name and column_name in df.columns:
                texts = df[column_name].dropna().tolist()
                return df, texts, column_name
            elif column_index < len(df.columns):
                column_name = df.columns[column_index]
                texts = df[column_name].dropna().tolist()
                return df, texts, column_name
            else:
                column_name = df.columns[0]
                texts = df[column_name].dropna().tolist()
                return df, texts, column_name
                
        except Exception as e:
            raise Exception(f"Ошибка чтения файла: {str(e)}")
    
    def analyze_batch(self, texts, progress_callback=None):
        results = []
        
        for i, text in enumerate(tqdm(texts, desc="Анализ текстов")):
            try:
                text = str(text) if pd.notna(text) else ""
                
                if len(text.strip()) < 3:
                    results.append({
                        'text': text,
                        'sentiment': 'neutral',
                        'topic_name': 'Другое',
                        'error': 'Текст слишком короткий'
                    })
                else:
                    sentiment_result = self.sentiment_analyzer.analyze(text)
                    topic_result = self.topic_classifier.classify(text)
                    
                    results.append({
                        'text': text,
                        'sentiment': sentiment_result['sentiment'],
                        'topic_name': topic_result['topic_name']
                    })
                
                if progress_callback:
                    progress_callback(i + 1, len(texts))
                    
            except Exception as e:
                print(f"Ошибка анализа текста {i}: {e}")
                results.append({
                    'text': text,
                    'sentiment': 'neutral',
                    'topic_name': 'Другое',
                    'error': str(e)
                })
        
        return results
    
    def create_result_dataframe(self, original_df, texts_column, results):
        """Создание результирующего DataFrame с тональностью и тематикой"""
        
        # Берем только проанализированные строки
        analyzed_count = len(results)
        result_df = original_df.iloc[:analyzed_count].copy()
        
        print(f"📊 Создание DataFrame: анализировано {analyzed_count} строк из {len(original_df)}")
        
        # Добавляем колонки с тональностью и тематикой
        result_df['Тональность'] = [r['sentiment'] for r in results]
        result_df['Тематика'] = [r['topic_name'] for r in results]
        
        return result_df
    
    def save_to_excel(self, df, original_filename, output_dir='downloads'):
        os.makedirs(output_dir, exist_ok=True)
        
        base_name = secure_filename(original_filename)
        name_without_ext = os.path.splitext(base_name)[0]
        output_filename = f"{name_without_ext}_analyzed_{int(time.time())}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Результаты анализа', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Результаты анализа']
            
            self._format_excel(worksheet, df)
        
        return output_path, output_filename
    
    def _format_excel(self, worksheet, df):
        """Форматирование Excel файла - компактный вариант"""
        
        # Более компактный шрифт для заголовков
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        
        # Тонкие границы
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Стиль для данных
        data_font = Font(size=9)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        
        # Форматируем заголовки
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Устанавливаем высоту строк по умолчанию
        worksheet.row_dimensions[1].height = 15
        
        # === НАСТРОЙКА ШИРИНЫ КОЛОНОК ===
        # Получаем буквы колонок
        col_letters = []
        for i in range(len(df.columns)):
            col_letters.append(openpyxl.utils.get_column_letter(i + 1))
    
        
        # Устанавливаем ширину для каждой колонки
        for col_idx, col_letter in enumerate(col_letters, 1):
            col_name = df.columns[col_idx-1] if col_idx-1 < len(df.columns) else ""
            
            # === НАСТРОЙКА ПО УМОЛЧАНИЮ ===
            default_width = 12  # Стандартная ширина
            
            # === ИНДИВИДУАЛЬНЫЕ НАСТРОЙКИ ===
            
     
            
            # Колонка D - делаем ШИРЕ (для текстов)
            if col_letter == 'D':
                worksheet.column_dimensions[col_letter].width = 30  # Широкая колонка для текстов
            
            # Колонки с тональностью (обычно идут после D)
            elif 'Тональность' in col_name:
                worksheet.column_dimensions[col_letter].width = 14
            
            # Колонки с тематикой
            elif 'Тематика' in col_name:
                worksheet.column_dimensions[col_letter].width = 18
            
            # Колонки с уверенностью
            elif 'Уверенность' in col_name:
                worksheet.column_dimensions[col_letter].width = 12
            
            # Колонка с эмодзи
            elif 'Эмодзи' in col_name:
                worksheet.column_dimensions[col_letter].width = 8
            
            # Альтернативные темы
            elif 'Альтернативные' in col_name:
                worksheet.column_dimensions[col_letter].width = 20
            
            # Все остальные колонки
            else:
                # Автоподбор ширины по содержимому, но с ограничением
                max_length = 0
                for row in worksheet.iter_rows(min_row=2, max_row=min(10, worksheet.max_row), 
                                            min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                width = min(max(max_length + 2, default_width), 18)
                worksheet.column_dimensions[col_letter].width = width
        
        # Форматируем данные
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
            worksheet.row_dimensions[row_idx].height = 13
            
            for col_idx, cell in enumerate(row, 1):
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Цвет для тональности
                if df.columns[col_idx-1] == 'Тональность':
                    sentiment = cell.value
                    if sentiment in self.sentiment_colors:
                        colors = {
                            'positive': 'E2F0D9',
                            'negative': 'FCE4D6',
                            'neutral': 'FFF2CC'
                        }
                        cell.fill = PatternFill(
                            start_color=colors.get(sentiment, 'FFFFFF'),
                            end_color=colors.get(sentiment, 'FFFFFF'),
                            fill_type="solid"
                        )
        
        # Добавляем фильтры
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Фиксируем шапку
        worksheet.freeze_panes = 'A2'
    
    print("✅ Excel отформатирован с индивидуальными настройками колонок")


class BatchAnalyzer:
    def __init__(self, excel_processor):
        self.excel_processor = excel_processor
        self.jobs = {}
        self.job_counter = 0
    
    def create_job(self, filepath, original_filename, column_info):
        job_id = f"job_{int(time.time())}_{self.job_counter}"
        self.job_counter += 1
        
        self.jobs[job_id] = {
            'id': job_id,
            'filepath': filepath,
            'original_filename': original_filename,
            'column_info': column_info,
            'status': 'pending',
            'progress': 0,
            'total': 0,
            'result_path': None,
            'result_filename': None,
            'error': None,
            'created_at': time.time()
        }
        
        return job_id
    
    def update_job_progress(self, job_id, current, total):
        if job_id in self.jobs:
            self.jobs[job_id]['progress'] = current
            self.jobs[job_id]['total'] = total
            self.jobs[job_id]['status'] = 'processing'
    
    def complete_job(self, job_id, result_path, result_filename):
        if job_id in self.jobs:
            self.jobs[job_id]['status'] = 'completed'
            self.jobs[job_id]['progress'] = self.jobs[job_id]['total']
            self.jobs[job_id]['result_path'] = result_path
            self.jobs[job_id]['result_filename'] = result_filename
    
    def fail_job(self, job_id, error):
        if job_id in self.jobs:
            self.jobs[job_id]['status'] = 'error'
            self.jobs[job_id]['error'] = str(error)
    
    def get_job(self, job_id):
        return self.jobs.get(job_id)
    
    def cleanup_old_jobs(self, max_age=3600):
        current_time = time.time()
        to_delete = []
        
        for job_id, job in self.jobs.items():
            if current_time - job['created_at'] > max_age:
                to_delete.append(job_id)
                
                if job.get('filepath') and os.path.exists(job['filepath']):
                    try:
                        os.remove(job['filepath'])
                    except:
                        pass
                
                if job.get('result_path') and os.path.exists(job['result_path']):
                    try:
                        os.remove(job['result_path'])
                    except:
                        pass
        
        for job_id in to_delete:
            del self.jobs[job_id]